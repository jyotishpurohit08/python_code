from selenium import webdriver
from selenium.webdriver.common.by import By
import openpyxl

# Initialize Selenium WebDriver
driver = webdriver.Firefox()
driver.maximize_window()
driver.get("https://www.amazon.in/")
driver.implicitly_wait(10)

# Search for Samsung phone
search_box = driver.find_element(By.XPATH, "//input[@id='twotabsearchtextbox']")
search_box.send_keys("Samsung phone")
search_button = driver.find_element(By.XPATH, "//*[@id='nav-search-submit-button']")
search_button.click()

# Filter for Samsung products
samsung_filter = driver.find_element(By.XPATH, "//span[text()='Samsung']")
samsung_filter.click()

# Extract product names
product_names = driver.find_elements(By.XPATH, "//span[contains(@class,'a-size-medium a-color-base a-text-normal')]")
product_names_list = [product.text for product in product_names]

# Extract product prices
product_prices = driver.find_elements(By.XPATH, "//span[contains(@class,'price-whole')]")
product_prices_list = [price.text for price in product_prices]

# Save data to Excel sheet
wb = openpyxl.Workbook()
sheet = wb.active

# Write headers
sheet['A1'] = 'Product Name'
sheet['B1'] = 'Price'

# Write data
for i in range(len(product_names_list)):
    sheet.cell(row=i+2, column=1).value = product_names_list[i]
    sheet.cell(row=i+2, column=2).value = product_prices_list[i]

# Save the Excel file
wb.save('amazon_products.xlsx')

# Close the browser
driver.quit()
